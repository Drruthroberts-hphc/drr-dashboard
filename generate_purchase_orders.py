"""
Purchase Order Workbook Generator
==================================
Generates a ready-to-action Excel workbook with reorder quantities
grouped by vendor and priority. Pulls live Shopify inventory data.

Usage:
    python3 generate_purchase_orders.py
"""

import sys
import os
import logging
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from collectors.inventory_collector import collect_inventory_data, MANUFACTURED_VENDORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styling ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='2D5F2D', end_color='2D5F2D', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill(start_color='4A7C59', end_color='4A7C59', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
ROW_ALT_FILL = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
URGENT_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
URGENT_FONT = Font(name='Calibri', bold=True, color='FFFFFF')
WARNING_FILL = PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid')
OOS_FILL = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')
OOS_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
OK_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
BLUE_FILL = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
SUBTOTAL_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
SUBTOTAL_FONT = Font(name='Calibri', bold=True, size=11)
GRAND_TOTAL_FILL = PatternFill(start_color='2D5F2D', end_color='2D5F2D', fill_type='solid')
GRAND_TOTAL_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
MONEY_FORMAT = '#,##0.00'
CHECK_FONT = Font(name='Calibri', size=14)

# ── Exclusions ────────────────────────────────────────────────────────────
# Products to exclude from purchase orders
EXCLUDE_VENDORS = {
    'Standard Process Inc',   # Drop-shipped, not stocked
}
EXCLUDE_PRODUCTS = {
    'Liposomal Glutathione 4 oz',  # Replaced by Earth Buddy version
}

# ── Vendor order info ─────────────────────────────────────────────────────
VENDOR_INFO = {
    'Real Mushrooms': {
        'portal': 'realmushrooms.com (Practitioner Portal)',
        'discount': '40% off retail (flat)',
        'shipping': 'Flat $10',
        'notes': 'ORDER ALREADY PLACED - see separate Real Mushrooms order',
    },
    'Microbiome Labs': {
        'portal': 'microbiomelabs.com/practitioner/ or Fullscript',
        'discount': 'Behind login - check both portals',
        'shipping': 'Check Fullscript vs direct',
        'notes': 'Compare Fullscript pricing to direct. FidoSpore, MegaSpore, RestorFlora all same vendor.',
    },
    'NOW': {
        'portal': 'Standard distributor channels',
        'discount': 'Distributor pricing',
        'shipping': 'Varies by distributor',
        'notes': 'Calcium Citrate and Kelp are high-volume CrockPET Diet staples.',
    },
    'Pure Encapsulations': {
        'portal': 'Fullscript',
        'discount': 'Fullscript practitioner pricing',
        'shipping': 'Fullscript standard',
        'notes': 'Available on Fullscript.',
    },
    'Glacier Peak Holistics': {
        'portal': 'glacierpeakholistics.com',
        'discount': 'Wholesale 6-packs available, pricing behind login',
        'shipping': 'Ask about wholesale shipping',
        'notes': 'Retail $129.95/test. Ask about 6-pack and 12-pack pricing.',
    },
    'Amber Naturalz': {
        'portal': 'Faire or ambernaturalz.com',
        'discount': 'Check Faire for new retailer deals',
        'shipping': 'Check Faire',
        'notes': 'HWF Clean Heart. Check if Faire has intro discount.',
    },
    'Herb Pharm': {
        'portal': 'herb-pharm.com/pages/wholesale',
        'discount': 'Behind login',
        'shipping': 'Free over $49',
        'notes': 'Wholesale account required. Free shipping over $49.',
    },
    'Allergy Research Group': {
        'portal': 'allergyresearchgroup.com or Fullscript',
        'discount': 'Practitioner wholesale + volume discounts',
        'shipping': 'Check both portals',
        'notes': 'Compare Fullscript vs direct practitioner pricing.',
    },
    'Jarrow Formulas': {
        'portal': 'jarrow.com or Fullscript',
        'discount': 'Reseller pricing',
        'shipping': 'Check Fullscript',
        'notes': 'Available on Fullscript. Compare pricing.',
    },
    'Vital Nutrients': {
        'portal': 'vitalnutrients.co or Fullscript',
        'discount': 'Practitioner volume discounts',
        'shipping': 'Check both',
        'notes': 'Quick practitioner application. Available on Fullscript.',
    },
    'FleasGone': {
        'portal': 'Contact directly',
        'discount': 'Volume deal - order larger qty',
        'shipping': 'Ask about bulk shipping',
        'notes': 'Plan larger volume order to negotiate better pricing.',
    },
    'Earth Buddy': {
        'portal': 'Contact directly',
        'discount': 'Unknown - inquire',
        'shipping': 'Unknown',
        'notes': 'Replacement for Empirical Labs Liposomal Glutathione.',
    },
    'Adored Beast': {
        'portal': 'Contact directly - no public wholesale program',
        'discount': 'Unknown',
        'shipping': 'Unknown',
        'notes': 'No formal wholesale program found. Contact directly.',
    },
    'Dr. Ruth Roberts': {
        'portal': 'Direct manufacturer',
        'discount': 'N/A - manufactured product',
        'shipping': 'N/A',
        'notes': 'Rush order: 100 each in ~3 weeks. Full production run in 4-6 weeks.',
    },
}

# Vendor wholesale cost multipliers (fraction of retail)
# 1.0 = retail, 0.6 = 40% off, 0.5 = 50% off
VENDOR_COST_MULTIPLIER = {
    'Real Mushrooms': 0.60,      # 40% off confirmed
    'Microbiome Labs': 0.65,     # Estimated practitioner
    'NOW': 0.55,                 # Distributor pricing
    'Pure Encapsulations': 0.65, # Fullscript
    'Glacier Peak Holistics': 0.65,
    'Amber Naturalz': 0.65,
    'Herb Pharm': 0.60,
    'Allergy Research Group': 0.60,
    'Jarrow Formulas': 0.60,
    'Vital Nutrients': 0.65,
    'FleasGone': 0.60,
    'Earth Buddy': 0.65,
    'Adored Beast': 0.65,
    'Dr. Ruth Roberts': 0.40,    # Manufacturing cost
}


def _apply_header(ws, headers):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = f'A{ws.max_row + 1}'


def _apply_border_row(ws, row, num_cols):
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER


def _auto_width(ws, min_w=10, max_w=45, **kwargs):
    # Accept max_width/min_width as aliases
    effective_max = kwargs.get('max_width', max_w)
    effective_min = kwargs.get('min_width', min_w)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, effective_min), effective_max)


def _build_order_items(items):
    """
    Build the purchase order list from inventory data.
    Returns dict of vendor -> list of order items, split by priority.
    """
    orders_this_week = []
    orders_next_week = []

    for item in items:
        if item['status'] != 'active':
            continue
        if item['vendor'] in EXCLUDE_VENDORS:
            continue
        if item['product_title'] in EXCLUDE_PRODUCTS:
            continue

        daily = item['avg_daily_sales']
        stock = item['inventory_quantity']
        sold = item['units_sold_90d']
        dos = item['days_of_stock']

        if sold == 0:
            continue  # No sales, no order needed

        # Calculate order qty: get to 90 days of supply
        target_days = 90
        if item['vendor'] in MANUFACTURED_VENDORS:
            target_days = 120  # Longer runway for manufactured

        target_stock = round(daily * target_days)
        order_qty = max(target_stock - stock, 0)
        if order_qty < 2 and sold > 0:
            order_qty = 2  # Minimum practical order

        # Special cases
        if item['vendor'] == 'FleasGone':
            order_qty = max(order_qty, 20)  # Larger volume for deal

        cost_mult = VENDOR_COST_MULTIPLIER.get(item['vendor'], 0.60)
        est_cost = round(order_qty * item['price'] * cost_mult, 2)

        order_item = {
            'vendor': item['vendor'],
            'product': item['product_title'],
            'sku': item['sku'],
            'current_stock': stock,
            'sold_90d': sold,
            'daily_rate': daily,
            'days_left': dos,
            'order_qty': order_qty,
            'retail_price': item['price'],
            'est_wholesale': round(item['price'] * cost_mult, 2),
            'est_total': est_cost,
            'urgency': item['urgency'],
        }

        # Priority split
        if isinstance(dos, (int, float)) and dos <= 90:
            orders_this_week.append(order_item)
        elif order_qty > 0:
            orders_next_week.append(order_item)

    # Sort by days left ascending within each list
    def sort_key(x):
        d = x['days_left']
        return d if isinstance(d, (int, float)) else 99999
    orders_this_week.sort(key=sort_key)
    orders_next_week.sort(key=sort_key)

    return orders_this_week, orders_next_week


def _write_purchase_orders(wb, this_week, next_week):
    """Tab 1: Purchase Orders by vendor, grouped and subtotaled."""
    ws = wb.active
    ws.title = 'Purchase Orders'

    headers = [
        'Ordered?', 'Vendor', 'Product', 'SKU',
        'Current Stock', 'Sold (90d)', 'Daily Rate', 'Days Left',
        'Order Qty', 'Est. Unit (wholesale)', 'Est. Line Total', 'Notes',
    ]
    num_cols = len(headers)

    # ── THIS WEEK section ──
    ws.append(['ORDER THIS WEEK'] + [''] * (num_cols - 1))
    section_row = ws.max_row
    ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=num_cols)
    cell = ws.cell(row=section_row, column=1)
    cell.fill = OOS_FILL
    cell.font = SECTION_FONT
    cell.alignment = CENTER_ALIGN

    _apply_header(ws, headers)

    # Group this_week by vendor
    vendors_this_week = {}
    for item in this_week:
        v = item['vendor']
        if v == 'Real Mushrooms':
            continue  # Already ordered
        vendors_this_week.setdefault(v, []).append(item)

    grand_total_this_week = 0

    for vendor in sorted(vendors_this_week.keys(), key=lambda v: min(
        i['days_left'] if isinstance(i['days_left'], (int, float)) else 999
        for i in vendors_this_week[v]
    )):
        items = vendors_this_week[vendor]
        vendor_total = 0

        for item in items:
            urgency_note = ''
            if item['urgency'] == 'OUT OF STOCK':
                urgency_note = 'OUT OF STOCK'
            elif item['urgency'] == 'URGENT - Order Now':
                urgency_note = 'URGENT'

            row_data = [
                '[ ]',
                item['vendor'],
                item['product'][:55],
                item['sku'],
                item['current_stock'],
                item['sold_90d'],
                round(item['daily_rate'], 2),
                item['days_left'] if isinstance(item['days_left'], (int, float)) else 'N/A',
                item['order_qty'],
                item['est_wholesale'],
                item['est_total'],
                urgency_note,
            ]
            ws.append(row_data)
            r = ws.max_row
            _apply_border_row(ws, r, num_cols)

            # Format money columns
            ws.cell(row=r, column=10).number_format = MONEY_FORMAT
            ws.cell(row=r, column=11).number_format = MONEY_FORMAT

            # Color urgency
            if item['urgency'] == 'OUT OF STOCK':
                for c in range(1, num_cols + 1):
                    ws.cell(row=r, column=c).fill = OOS_FILL
                    ws.cell(row=r, column=c).font = OOS_FONT
            elif item['urgency'] in ('URGENT - Order Now', 'Reorder Soon'):
                for c in range(1, num_cols + 1):
                    ws.cell(row=r, column=c).fill = URGENT_FILL

            # Checkbox column
            ws.cell(row=r, column=1).font = CHECK_FONT
            ws.cell(row=r, column=1).alignment = CENTER_ALIGN

            vendor_total += item['est_total']

        # Vendor subtotal row
        ws.append(['', f'{vendor} Subtotal', '', '', '', '', '', '',
                    sum(i['order_qty'] for i in items), '', vendor_total, ''])
        r = ws.max_row
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
            ws.cell(row=r, column=c).font = SUBTOTAL_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=11).number_format = MONEY_FORMAT

        grand_total_this_week += vendor_total

    # Grand total this week
    ws.append([''])
    ws.append(['', 'THIS WEEK TOTAL', '', '', '', '', '', '', '', '', grand_total_this_week, ''])
    r = ws.max_row
    for c in range(1, num_cols + 1):
        ws.cell(row=r, column=c).fill = GRAND_TOTAL_FILL
        ws.cell(row=r, column=c).font = GRAND_TOTAL_FONT
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=11).number_format = MONEY_FORMAT

    # ── Blank separator ──
    ws.append([''])
    ws.append([''])

    # ── NEXT WEEK section ──
    ws.append(['ORDER NEXT WEEK'] + [''] * (num_cols - 1))
    section_row = ws.max_row
    ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=num_cols)
    cell = ws.cell(row=section_row, column=1)
    cell.fill = WARNING_FILL
    cell.font = SECTION_FONT
    cell.alignment = CENTER_ALIGN

    _apply_header(ws, headers)

    # Group next_week by vendor
    vendors_next_week = {}
    for item in next_week:
        v = item['vendor']
        if v == 'Real Mushrooms':
            continue
        vendors_next_week.setdefault(v, []).append(item)

    grand_total_next_week = 0

    for vendor in sorted(vendors_next_week.keys()):
        items = vendors_next_week[vendor]
        vendor_total = 0

        for item in items:
            row_data = [
                '[ ]',
                item['vendor'],
                item['product'][:55],
                item['sku'],
                item['current_stock'],
                item['sold_90d'],
                round(item['daily_rate'], 2),
                item['days_left'] if isinstance(item['days_left'], (int, float)) else 'N/A',
                item['order_qty'],
                item['est_wholesale'],
                item['est_total'],
                '',
            ]
            ws.append(row_data)
            r = ws.max_row
            _apply_border_row(ws, r, num_cols)
            ws.cell(row=r, column=10).number_format = MONEY_FORMAT
            ws.cell(row=r, column=11).number_format = MONEY_FORMAT
            ws.cell(row=r, column=1).font = CHECK_FONT
            ws.cell(row=r, column=1).alignment = CENTER_ALIGN

            vendor_total += item['est_total']

        # Vendor subtotal
        ws.append(['', f'{vendor} Subtotal', '', '', '', '', '', '',
                    sum(i['order_qty'] for i in items), '', vendor_total, ''])
        r = ws.max_row
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
            ws.cell(row=r, column=c).font = SUBTOTAL_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=11).number_format = MONEY_FORMAT

        grand_total_next_week += vendor_total

    # Grand total next week
    ws.append([''])
    ws.append(['', 'NEXT WEEK TOTAL', '', '', '', '', '', '', '', '', grand_total_next_week, ''])
    r = ws.max_row
    for c in range(1, num_cols + 1):
        ws.cell(row=r, column=c).fill = GRAND_TOTAL_FILL
        ws.cell(row=r, column=c).font = GRAND_TOTAL_FONT
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=11).number_format = MONEY_FORMAT

    # ── Combined total ──
    ws.append([''])
    ws.append(['', 'COMBINED TOTAL (Both Weeks)', '', '', '', '', '', '', '', '',
               grand_total_this_week + grand_total_next_week, ''])
    r = ws.max_row
    for c in range(1, num_cols + 1):
        ws.cell(row=r, column=c).fill = PatternFill(start_color='1A3A1A', end_color='1A3A1A', fill_type='solid')
        ws.cell(row=r, column=c).font = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=11).number_format = MONEY_FORMAT

    _auto_width(ws)
    # Widen key columns
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['L'].width = 20


def _write_manufactured_orders(wb, items):
    """Tab 2: Manufactured product orders (Dr. Ruth Roberts)."""
    ws = wb.create_sheet('Manufactured Products')

    headers = ['Product', 'Current Stock', 'Daily Rate', 'Days Left',
               'Rush Qty (3 wks)', 'Stock After Rush', 'Runway After Rush',
               'Full Run Qty (4-6 wks)', 'Notes']
    num_cols = len(headers)

    _apply_header(ws, headers)

    manufactured = [i for i in items if i['vendor'] in MANUFACTURED_VENDORS
                    and i['status'] == 'active' and i['units_sold_90d'] > 0]
    manufactured.sort(key=lambda x: x['days_of_stock'] if isinstance(x['days_of_stock'], (int, float)) else 99999)

    for item in manufactured:
        daily = item['avg_daily_sales']
        stock = item['inventory_quantity']
        dos = item['days_of_stock']

        rush_qty = 100
        burn_in_21d = round(daily * 21)
        stock_after_rush = stock - burn_in_21d + rush_qty
        runway_after = round(stock_after_rush / daily) if daily > 0 else 'N/A'

        # Full run: enough for 120 days minus what we'll have after rush
        full_target = round(daily * 120)
        full_qty = max(full_target - stock_after_rush, 0)
        # Round up to nearest 50
        full_qty = ((full_qty + 49) // 50) * 50

        notes = ''
        if isinstance(dos, (int, float)) and dos < 84:
            notes = 'WILL STOCK OUT BEFORE DELIVERY without rush'

        ws.append([
            item['product_title'], stock, round(daily, 2), dos,
            rush_qty, stock_after_rush, runway_after,
            full_qty, notes,
        ])
        r = ws.max_row
        _apply_border_row(ws, r, num_cols)

        if isinstance(dos, (int, float)) and dos < 84:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = URGENT_FILL

    _auto_width(ws, max_width=55)


def _write_vendor_contacts(wb):
    """Tab 3: Vendor contact info and ordering portals."""
    ws = wb.create_sheet('Vendor Portals & Notes')

    headers = ['Vendor', 'Order Portal', 'Discount', 'Shipping', 'Key Notes']
    num_cols = len(headers)

    _apply_header(ws, headers)

    for vendor in sorted(VENDOR_INFO.keys()):
        info = VENDOR_INFO[vendor]
        ws.append([
            vendor,
            info.get('portal', ''),
            info.get('discount', ''),
            info.get('shipping', ''),
            info.get('notes', ''),
        ])
        r = ws.max_row
        _apply_border_row(ws, r, num_cols)
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).alignment = WRAP_ALIGN
        if r % 2 == 0:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = ROW_ALT_FILL

    _auto_width(ws, max_width=55)


def _write_vendor_questions(wb):
    """Tab 4: Questions for inventory person to ask each vendor."""
    ws = wb.create_sheet('Questions for Vendors')

    questions = [
        ('Pricing', 'What are your wholesale discount tiers? (e.g. 5+/10+/case)', 'Higher volume = lower cost per unit'),
        ('Pricing', 'Do you offer introductory/first-order discounts?', 'One-time savings on initial order'),
        ('Pricing', 'Is there a MAP (Minimum Advertised Price) policy?', 'Determines lowest price you can list on site'),
        ('Shipping', 'What is your free shipping threshold?', 'Optimize order size to avoid shipping cost'),
        ('Shipping', 'Standard shipping cost under the free threshold?', 'Factor into cost-per-unit calculations'),
        ('Shipping', 'Typical lead time from order to delivery?', 'Critical for accurate reorder point timing'),
        ('Orders', 'Minimum order quantity or dollar amount?', 'Some vendors require minimums per SKU or per order'),
        ('Orders', 'Auto-ship / standing order discounts?', 'Additional % off for recurring orders'),
        ('Orders', 'Net-30 or Net-60 payment terms available?', 'Cash flow: sell product before paying for it'),
        ('Promos', 'Seasonal promotions or buying windows?', 'Time purchases for best pricing'),
        ('Promos', 'Loyalty or volume rewards program?', 'Long-term savings that compound over time'),
        ('Account', 'Wholesale vs. retail price difference?', 'Target: 30-50% off retail'),
        ('Account', 'Practitioner referral/commission program?', 'Earn commission on patient direct purchases'),
        ('Product', 'Shelf life? Can you request longer-dated stock?', 'Avoid expiring inventory sitting on shelves'),
        ('Product', 'Return/exchange policy for wholesale?', 'Know options if product arrives damaged or doesnt sell'),
    ]

    headers = ['Category', 'Question to Ask', 'Why This Matters']
    num_cols = len(headers)
    _apply_header(ws, headers)

    for cat, question, why in questions:
        ws.append([cat, question, why])
        r = ws.max_row
        _apply_border_row(ws, r, num_cols)
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).alignment = WRAP_ALIGN
        if r % 2 == 0:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = ROW_ALT_FILL

    _auto_width(ws, max_width=60)


def generate_purchase_orders():
    """Main entry: pull Shopify data and generate purchase order workbook."""
    logger.info("Generating purchase order workbook...")

    data = collect_inventory_data()
    this_week, next_week = _build_order_items(data['items'])

    wb = Workbook()

    _write_purchase_orders(wb, this_week, next_week)
    _write_manufactured_orders(wb, data['items'])
    _write_vendor_contacts(wb)
    _write_vendor_questions(wb)

    today = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.expanduser('~/Desktop/Code output')
    path = os.path.join(output_dir, f'Purchase_Orders_{today}.xlsx')

    wb.save(path)
    logger.info(f"Purchase order workbook saved: {path}")

    # Summary
    tw_total = sum(i['est_total'] for i in this_week if i['vendor'] != 'Real Mushrooms')
    nw_total = sum(i['est_total'] for i in next_week if i['vendor'] != 'Real Mushrooms')
    tw_vendors = set(i['vendor'] for i in this_week if i['vendor'] != 'Real Mushrooms')
    nw_vendors = set(i['vendor'] for i in next_week if i['vendor'] != 'Real Mushrooms')

    print(f"\nPurchase Order Workbook: {path}")
    print(f"\nThis Week: {len(tw_vendors)} vendors, est. ${tw_total:,.2f}")
    for v in sorted(tw_vendors):
        v_items = [i for i in this_week if i['vendor'] == v]
        v_total = sum(i['est_total'] for i in v_items)
        print(f"  {v}: {len(v_items)} items, {sum(i['order_qty'] for i in v_items)} units, ${v_total:,.2f}")

    print(f"\nNext Week: {len(nw_vendors)} vendors, est. ${nw_total:,.2f}")
    for v in sorted(nw_vendors):
        v_items = [i for i in next_week if i['vendor'] == v]
        v_total = sum(i['est_total'] for i in v_items)
        print(f"  {v}: {len(v_items)} items, {sum(i['order_qty'] for i in v_items)} units, ${v_total:,.2f}")

    print(f"\nCombined: ${tw_total + nw_total:,.2f}")

    return path


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
    generate_purchase_orders()
