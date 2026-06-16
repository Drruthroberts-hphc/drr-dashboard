"""
Weekly Inventory Reorder Report Generator
==========================================
Pulls live data from Shopify, calculates reorder points, and outputs
a formatted Excel workbook to the Desktop.

Usage:
    python3 generate_inventory_report.py

Output:
    ~/Desktop/Code output/Inventory_Reorder_Guide_YYYY-MM-DD.xlsx
"""

import sys
import os
import logging
from datetime import datetime

# Add parent dir so we can import collectors and config
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from collectors.inventory_collector import collect_inventory_data
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styling ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='2D5F2D', end_color='2D5F2D', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
ROW_ALT_FILL = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
URGENT_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid')
OOS_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
OOS_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
OK_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# Vendor wholesale terms (static reference data)
VENDOR_TERMS = [
    {
        'vendor': 'AnimalBiome',
        'account_type': 'Wholesale (existing account)',
        'discount': 'Tiered: $0-$999 / $1,000-$2,499 (exact % behind login)',
        'free_shipping': 'Retail: free >$50 / Wholesale: unknown',
        'min_order': 'Unknown',
        'fullscript': 'No',
        'faire': 'No',
        'notes': 'MAP: no advertising below 30% off MSRP. Portal: animalbiome.biz',
        'action': 'Log in to animalbiome.biz to view tier pricing. Ask about $2,500+ tiers.',
    },
    {
        'vendor': 'Empirical Labs',
        'account_type': 'Practitioner account',
        'discount': 'Not publicly disclosed',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'No',
        'faire': 'No',
        'notes': 'Available via Pro Supplement Center, OHP Health',
        'action': 'Contact directly for wholesale pricing',
    },
    {
        'vendor': 'Allergy Research Group',
        'account_type': 'Free Practitioner Account (credentials required)',
        'discount': 'Wholesale + volume qty discounts',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'YES',
        'faire': 'No',
        'notes': 'Patient Participation Program: patients get discount, you earn commission',
        'action': 'Apply at allergyresearchgroup.com. Compare Fullscript pricing.',
    },
    {
        'vendor': 'Herb Pharm',
        'account_type': 'Wholesale + Professional tiers',
        'discount': 'Not publicly disclosed',
        'free_shipping': 'Free over $49 (after discounts)',
        'min_order': 'None stated (max $1,000 online)',
        'fullscript': 'No',
        'faire': 'No',
        'notes': 'Orders >$1,000: call 800-348-4372',
        'action': 'Create wholesale account at herb-pharm.com',
    },
    {
        'vendor': 'Herbsmith',
        'account_type': 'Wholesale signup (no vet auth for supplements)',
        'discount': 'Not publicly disclosed',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'No',
        'faire': 'YES',
        'notes': 'Also on Wholesale Pet and Vetcove',
        'action': 'Check Faire for new retailer discounts',
    },
    {
        'vendor': 'Jarrow Formulas',
        'account_type': 'Reseller Application (license required)',
        'discount': 'Not publicly disclosed',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'YES',
        'faire': 'No',
        'notes': 'Standard supplement distributor',
        'action': 'Apply at jarrow.com. Compare Fullscript pricing.',
    },
    {
        'vendor': 'Vital Nutrients',
        'account_type': 'Practitioner Program (quick app)',
        'discount': 'Volume discounts (behind login)',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'YES',
        'faire': 'No',
        'notes': 'Quick application process',
        'action': 'Apply at vitalnutrients.co. Compare Fullscript pricing.',
    },
    {
        'vendor': 'Adored Beast',
        'account_type': 'No public wholesale program found',
        'discount': 'Unknown',
        'free_shipping': 'Unknown',
        'min_order': 'Unknown',
        'fullscript': 'No',
        'faire': 'No',
        'notes': 'Sold through independent retailers',
        'action': 'Contact directly at adoredbeast.com',
    },
    {
        'vendor': 'Real Mushrooms / 5 Defenders',
        'account_type': 'Practitioner account',
        'discount': '40% flat discount for ALL practitioners',
        'free_shipping': 'Flat-rate $10 (no free option)',
        'min_order': 'Not disclosed',
        'fullscript': 'No',
        'faire': 'No',
        'notes': '1,198+ practitioners. Discount NOT combinable with promos.',
        'action': 'Apply at realmushrooms.com. Bulk order to offset $10 shipping.',
    },
    {
        'vendor': 'MicroBiome Labs',
        'account_type': 'Practitioner account required',
        'discount': 'Behind login',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'YES',
        'faire': 'No',
        'notes': 'Products: FidoSpore, MegaSporeBiotic, RestorFlora',
        'action': 'Apply at microbiomelabs.com/practitioner/. Check Fullscript.',
    },
    {
        'vendor': 'Glacier Peak Holistics',
        'account_type': 'Wholesale available',
        'discount': '6-pack wholesale (pricing not public)',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'No',
        'faire': 'No',
        'notes': 'Retail: $89/test ($99 rush)',
        'action': 'Contact for wholesale 6-pack pricing',
    },
    {
        'vendor': 'UCARI',
        'account_type': 'Wholesale inquiry form',
        'discount': '10% off initial stocking + 1 free kit',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'No',
        'faire': 'No',
        'notes': 'In 2,000+ retail stores. 10-packs available.',
        'action': 'Submit inquiry at ucari.com/pages/wholesale-enquiry',
    },
    {
        'vendor': 'HWFA / Amber Naturalz',
        'account_type': 'Available on Faire',
        'discount': 'Not publicly disclosed',
        'free_shipping': 'Not disclosed',
        'min_order': 'Not disclosed',
        'fullscript': 'No',
        'faire': 'YES',
        'notes': 'Various authorized retailers',
        'action': 'Check Faire pricing. Contact Amber Naturalz directly.',
    },
]

VENDOR_QUESTIONS = [
    ('Pricing', 'What are your wholesale discount tiers? (e.g. 5+/10+/case)', 'Higher volume = lower cost. Need breakpoints to optimize orders.', 'Wholesale portal > Pricing page'),
    ('Pricing', 'Do you offer introductory/first-order discounts?', 'One-time savings. Take advantage before it expires.', 'Welcome emails or ask sales rep'),
    ('Pricing', 'Is there a MAP (Minimum Advertised Price) policy?', 'Determines lowest price you can list. Violating MAP can get you terminated.', 'Terms of service or MAP policy doc'),
    ('Shipping', 'What is your free shipping threshold?', 'Optimize order size to avoid shipping eating margins.', 'Checkout page or FAQ'),
    ('Shipping', 'Standard shipping cost under the free threshold?', 'Factor shipping into cost-per-unit calculations.', 'Checkout page or shipping calculator'),
    ('Shipping', 'Typical lead time from order to delivery?', 'Critical for accurate reorder points. Default is 14 days.', 'FAQ or track actual delivery times'),
    ('Orders', 'Minimum order quantity or dollar amount?', 'Some vendors require minimums per SKU or per order.', 'Terms of service or wholesale agreement'),
    ('Orders', 'Auto-ship / standing order discounts available?', 'Additional % off for recurring orders saves money.', 'Account settings or ask sales rep'),
    ('Orders', 'Is Net-30 or Net-60 payment available?', 'Cash flow: sell product before paying for it.', 'Payment terms in wholesale agreement'),
    ('Promotions', 'Seasonal promotions or buying windows?', 'Time purchases for best pricing.', 'Email newsletters or portal announcements'),
    ('Promotions', 'Loyalty or volume-based rewards program?', 'Long-term savings that compound over time.', 'Account dashboard or ask sales rep'),
    ('Account', 'Wholesale vs. retail price difference (% off)?', 'Quick margin comparison. Target: 30-50% off retail.', 'Compare wholesale to consumer site'),
    ('Account', 'Practitioner referral/commission program?', 'Earn commission on patient direct purchases.', '"Patient Programs" or "Referral" in dashboard'),
    ('Returns', 'Return/exchange policy for wholesale orders?', 'Know options if product arrives damaged or does not sell.', 'Terms of service'),
    ('Returns', 'Shelf life? Can you request longer-dated stock?', 'Avoid expiring inventory sitting on shelves.', 'Ask when ordering or check labels'),
]


def _apply_header_style(ws, num_cols):
    """Apply header styling and freeze top row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = 'A2'


def _auto_width(ws, min_width=10, max_width=45):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_tab_1_reorder(wb, items):
    """Tab 1: Active Inventory Reorder Status (main tab, refreshed weekly)."""
    ws = wb.active
    ws.title = 'Reorder Status'

    headers = [
        'Urgency', 'Product', 'Variant', 'SKU', 'Vendor',
        'Current Stock', 'Units Sold (90d)', 'Avg Daily Sales',
        'Lead Time (days)', 'Lead Time Demand', 'Safety Stock',
        'Reorder Point', 'Suggested Order Qty', 'Days of Stock Left',
        'Est. Order Cost (wholesale)', 'Retail Price',
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    active_items = [i for i in items if i['status'] == 'active']

    for idx, item in enumerate(active_items, start=2):
        row = [
            item['urgency'],
            item['product_title'],
            item['variant_title'] if item['variant_title'] != 'Default Title' else '',
            item['sku'],
            item['vendor'],
            item['inventory_quantity'],
            item['units_sold_90d'],
            item['avg_daily_sales'],
            item['lead_time_days'],
            item['lead_time_demand'],
            item['safety_stock'],
            item['reorder_point'],
            item['suggested_order_qty'],
            item['days_of_stock'],
            item['estimated_order_cost'],
            item['price'],
        ]
        ws.append(row)

        # Apply urgency coloring
        urgency = item['urgency']
        if urgency == 'OUT OF STOCK':
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=idx, column=col_idx)
                cell.fill = OOS_FILL
                cell.font = OOS_FONT
        elif urgency == 'URGENT - Order Now':
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col_idx).fill = URGENT_FILL
        elif urgency == 'Reorder Soon':
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col_idx).fill = WARNING_FILL
        elif urgency == 'OK':
            ws.cell(row=idx, column=1).fill = OK_FILL
        elif idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col_idx).fill = ROW_ALT_FILL

        # Borders
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col_idx).border = THIN_BORDER

    _auto_width(ws)


def _write_tab_2_discontinued(wb, items):
    """Tab 2: Discontinued / No-Sales items for review."""
    ws = wb.create_sheet('Discontinued & No Sales')

    headers = ['Product', 'SKU', 'Vendor', 'Status', 'Current Stock',
               'Units Sold (90d)', 'Retail Price', 'Action Needed']
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    no_sales = [i for i in items if i['urgency'] == 'No Sales' or i['status'] != 'active']

    for idx, item in enumerate(no_sales, start=2):
        action = ''
        if item['status'] != 'active':
            action = 'Archived/Draft - review'
        elif item['inventory_quantity'] > 0 and item['units_sold_90d'] == 0:
            action = 'Dead stock - consider clearance or discontinue'
        elif item['inventory_quantity'] == 0 and item['units_sold_90d'] == 0:
            action = 'Zero stock, zero sales - discontinue?'

        ws.append([
            item['product_title'], item['sku'], item['vendor'],
            item['status'], item['inventory_quantity'],
            item['units_sold_90d'], item['price'], action,
        ])

        if item['inventory_quantity'] > 0 and item['units_sold_90d'] == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col_idx).fill = YELLOW_FILL

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col_idx).border = THIN_BORDER

        if idx % 2 == 0 and not (item['inventory_quantity'] > 0 and item['units_sold_90d'] == 0):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col_idx).fill = ROW_ALT_FILL

    _auto_width(ws)


def _write_tab_3_vendor_terms(wb):
    """Tab 3: Vendor Wholesale Terms reference."""
    ws = wb.create_sheet('Vendor Wholesale Terms')

    headers = ['Vendor', 'Account Type', 'Discount/Pricing', 'Free Shipping',
               'Min Order', 'Fullscript?', 'Faire?', 'Key Notes', 'Action Items']
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    for idx, v in enumerate(VENDOR_TERMS, start=2):
        ws.append([
            v['vendor'], v['account_type'], v['discount'],
            v['free_shipping'], v['min_order'], v['fullscript'],
            v['faire'], v['notes'], v['action'],
        ])

        if v['fullscript'] == 'YES':
            ws.cell(row=idx, column=6).fill = OK_FILL
        if v['faire'] == 'YES':
            ws.cell(row=idx, column=7).fill = OK_FILL

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col_idx).border = THIN_BORDER
            ws.cell(row=idx, column=col_idx).alignment = WRAP_ALIGN

        if idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=idx, column=col_idx)
                if not c.fill or c.fill.start_color.rgb in ('00000000', 'FFFFFF'):
                    c.fill = ROW_ALT_FILL

    _auto_width(ws, max_width=50)


def _write_tab_4_questions(wb):
    """Tab 4: Vendor Questions Checklist for inventory person."""
    ws = wb.create_sheet('Vendor Questions Checklist')

    headers = ['Category', 'Question to Ask', 'Why This Matters', 'Where to Find It']
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    for idx, (cat, question, why, where) in enumerate(VENDOR_QUESTIONS, start=2):
        ws.append([cat, question, why, where])

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN

        if idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=idx, column=col_idx).fill = ROW_ALT_FILL

    _auto_width(ws, max_width=55)


def _write_tab_5_summary(wb, summary):
    """Tab 5: Collection summary / metadata."""
    ws = wb.create_sheet('Report Summary')

    headers = ['Metric', 'Value']
    ws.append(headers)
    _apply_header_style(ws, 2)

    rows = [
        ('Report Generated', summary['collection_date']),
        ('Sales Lookback Period', f"{summary['lookback_days']} days"),
        ('Default Lead Time', f"{summary['lead_time_days']} days"),
        ('Total Tracked Variants', summary['total_tracked_variants']),
        ('Active Variants', summary['active_variants']),
        ('URGENT Reorder Count', summary['urgent_reorder_count']),
        ('Reorder Soon Count', summary['reorder_soon_count']),
        ('Est. Total Order Cost (wholesale)', f"${summary['total_estimated_order_cost']:,.2f}"),
        ('', ''),
        ('NOTES', ''),
        ('Lead Time', 'Default 14 days. Update per vendor as you learn actual times.'),
        ('Safety Stock', '20% buffer above lead time demand. Adjust for seasonal items.'),
        ('Suggested Qty', 'Based on 60-day supply. Adjust to meet vendor minimums/free shipping.'),
        ('Est. Cost', 'Rough estimate at 60% of retail. Update with actual wholesale prices.'),
    ]

    for idx, (metric, value) in enumerate(rows, start=2):
        ws.append([metric, value])
        for col_idx in range(1, 3):
            ws.cell(row=idx, column=col_idx).border = THIN_BORDER

    _auto_width(ws, max_width=70)


def generate_report():
    """Pull live Shopify data and generate the Excel report."""
    logger.info("Generating weekly inventory reorder report...")

    data = collect_inventory_data()

    wb = Workbook()

    _write_tab_1_reorder(wb, data['items'])
    _write_tab_2_discontinued(wb, data['items'])
    _write_tab_3_vendor_terms(wb)
    _write_tab_4_questions(wb)
    _write_tab_5_summary(wb, data['summary'])

    # Save with date stamp
    today = datetime.now().strftime('%Y-%m-%d')
    output_dir = os.path.expanduser('~/Desktop/Code output')
    os.makedirs(output_dir, exist_ok=True)

    # Save both a dated version and a "latest" version
    dated_path = os.path.join(output_dir, f'Inventory_Reorder_Guide_{today}.xlsx')
    latest_path = os.path.join(output_dir, 'Inventory_Reorder_Guide_LATEST.xlsx')

    wb.save(dated_path)
    wb.save(latest_path)

    logger.info(f"Report saved to: {dated_path}")
    logger.info(f"Latest copy at: {latest_path}")

    return {
        'dated_path': dated_path,
        'latest_path': latest_path,
        'summary': data['summary'],
    }


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
    result = generate_report()
    print(f"\nReport generated successfully!")
    print(f"  Dated:  {result['dated_path']}")
    print(f"  Latest: {result['latest_path']}")
    print(f"\nSummary:")
    for k, v in result['summary'].items():
        print(f"  {k}: {v}")
